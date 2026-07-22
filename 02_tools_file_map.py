"""
FILE SYSTEM MAPPING TOOL
========================
ER 20.11.2024

This script creates a complete inventory of all files in a directory,
generating a detailed Excel report with comprehensive file information.

Features:
- Full recursive directory scanning
- Excel report including:
  - Relative folder path
  - Filename with hyperlink
  - File size in bytes
  - Last modified date
  - Creation date
- Error handling for inaccessible files
- Useful for:
  - File inventory management
  - Storage space analysis
  - File change tracking

Required libraries:
- os (standard library)
- openpyxl (install with: pip install openpyxl)
- time (standard library)

Usage:
- Modify the root_folder variable with the path to scan
- Run the script to generate the report
"""
import os
import openpyxl
from openpyxl.styles import Alignment
import time

def get_file_info(file_path):
    try:
        size = os.path.getsize(file_path)
        mod_time = os.path.getmtime(file_path)
        create_time = os.path.getctime(file_path)
        return size, mod_time, create_time
    except Exception as e:
        print(f"Errore nell'accesso al file {file_path}: {e}")
        return None, None, None

def create_excel_report(root_folder, output_file):
    # Print the root folder and output file path
    print(f"Root folder: {root_folder}")
    print(f"Output file: {output_file}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Report"

    headers = ["Folder Path", "File Name", "Size (Bytes)", "Last Modified", "Creation Date"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).alignment = Alignment(horizontal="center")

    row = 2  # Start from row 2 for data

    def explore_folder(current_folder, depth=0):
        nonlocal row
        for root, dirs, files in os.walk(current_folder):
            for file in files:
                file_path = os.path.join(root, file)
                size, mod_time, create_time = get_file_info(file_path)

                if size is not None:
                    mod_time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mod_time))
                    create_time_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(create_time))

                    folder_relative_path = os.path.relpath(root, start=root_folder)
                    file_name = file
                    hyperlink = f"'{file_path}'"  # Make sure the hyperlink path is correct

                    # Write data to the sheet
                    ws.cell(row=row, column=1, value=folder_relative_path)
                    ws.cell(row=row, column=2, value=file_name).hyperlink = hyperlink
                    ws.cell(row=row, column=3, value=size)
                    ws.cell(row=row, column=4, value=mod_time_str)
                    ws.cell(row=row, column=5, value=create_time_str)

                    row += 1

    explore_folder(root_folder)

    # Concatenate root_folder path with the file name and ensure it has the .xlsx extension
    output_file_path = root_folder + "\\" + "_FILE_MAPPING.xlsx"

    # Save the workbook
    wb.save(output_file_path)
    print(f"Report saved as {output_file_path}")

root_folder = r"C:\user\path"  # Path to your folder
output_file = "_FILE_MAPPING.xlsx"  # Output file name
create_excel_report(root_folder, output_file)
